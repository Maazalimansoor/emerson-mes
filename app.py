from flask import Flask, render_template, request, jsonify
import os
import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# ================= CONFIG =================
BASE_FOLDER = r"\\bt\BarTenderFiles\Traceability"
EXCEL_FILE = os.path.join(BASE_FOLDER, "Emerson_label.xlsx")

GLASS_PRINT_URL = "http://bt.corp.alpha-measure.com/bartender/Print/8ee74183-83ea-40a4-be49-ea5d0b988838/Glass%20Shop/Emerson%20Glass%20label%2033%20.btw"
QC_PRINT_URL = "http://bt.corp.alpha-measure.com/bartender/Print/13baa022-4b55-4a1c-b999-bacf11914079/Emerson/Emerson%20label%2033%20per%20page%20.btw#"

# ================= MASTER DATA =================
OPERATORS = {
    "Tamika Jones": "01",
    "Quy Don Le": "02",
    "Son Ngo": "03",
    "Huy Nguyen": "04",
    "Thanh N Nguyen": "05",
    "Binh Phan": "06",
    "Hoang tran": "07",
    "Cuong Vu": "08",
    "Julin Wu": "09",
}

PARTS = [
    "00310-0390-HGP1",
    "00310-0396-HGP1",
    "00310-0350-T001",
    "00310-0350-HHT1"
]

# ================= INIT EXCEL =================
def init_excel():
    os.makedirs(BASE_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Glass Label"
    ws1.append([
        "WO","Part","Stem","Crucible","Powder",
        "Name","OP No","Qty","Date"
    ])

    ws2 = wb.create_sheet("QC Label")
    ws2.append([
        "Part Number",
        "Glass Lot ID",
        "Date"
    ])

    wb.save(EXCEL_FILE)
    wb.close()

# ================= SAVE GLASS =================
def save_glass(d):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Glass Label"]

    ws.append([
        d["wo"],
        d["part"],
        d["stem"],
        d["crucible"],
        d["powder"],
        d["name"],
        d["opno"],
        d["qty"],
        datetime.date.today().strftime("%m%d%y")
    ])

    wb.save(EXCEL_FILE)
    wb.close()

# ================= SAVE QC =================
def save_qc(d):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["QC Label"]

    ws.append([
        d["part"],
        d["lot"],
        datetime.date.today().strftime("%m%d%y")
    ])

    wb.save(EXCEL_FILE)
    wb.close()

# ================= ROUTES =================
@app.route("/")
def index():
    return render_template(
        "index.html",
        operators=OPERATORS,
        parts=PARTS,
        glass_url=GLASS_PRINT_URL,
        qc_url=QC_PRINT_URL
    )

@app.route("/save_glass", methods=["POST"])
def glass():
    d = request.json

    required = ["wo","part","stem","crucible","powder","name","opno","qty"]

    for k in required:
        if not str(d.get(k,"")).strip():
            return jsonify({"msg": f"Missing field: {k}", "status": "error"})

    if not d["qty"].isdigit():
        return jsonify({"msg": "Qty must be numeric", "status": "error"})

    save_glass(d)

    return jsonify({"msg": "Glass Saved Successfully", "status": "ok"})

@app.route("/save_qc", methods=["POST"])
def qc():
    d = request.json

    if not d.get("part") or not d.get("lot"):
        return jsonify({"msg": "Missing QC data", "status": "error"})

    save_qc(d)

    return jsonify({"msg": "QC Saved Successfully", "status": "ok"})

# ================= START =================
if __name__ == "__main__":
    init_excel()
    app.run(host="0.0.0.0", port=5000, debug=False)